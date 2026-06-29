from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "content" / "portfolio-v3-content-tracker.xlsx"

PAGES = {
    "0": ("Header", "Global navigation and identity shown on every page."),
    "1": ("Profile / Home", "Opening profile, portrait, credentials and core capabilities."),
    "2": ("Strategy", "Six-part strategic approach or working philosophy."),
    "3": ("Selected Work", "Project timeline, case-study cards and project-detail panel."),
    "4": ("About / Resume / Contact", "Background, education, supporting information and contact details."),
    "5": ("Footer", "Closing statement and copyright information."),
}

rows = []


def add(code, page, section, purpose, key, placeholder, value_type="Text"):
    rows.append(
        {
            "Hierarchy ID": code,
            "Page": page,
            "Page Name": PAGES[page][0],
            "Section": section,
            "Field Purpose": purpose,
            "Technical Key": key,
            "Placeholder / Guidance": placeholder,
            "Value Type": value_type,
            "Your Value": "",
            "Status": "pending",
            "Notes": "",
        }
    )


# Page 0 — Header
# Brand initials and fixed navigation labels are structural in this concept,
# so they are intentionally excluded from the fillable tracker.
add("0.3", "0", "Header action", "Resume button label", "header_resume_label", "Example: Resume")

# Page 1 — Profile / Home
add("1.1", "1", "Hero", "Introductory greeting", "hero_eyebrow", "Example: Hello, I’m")
add("1.2", "1", "Hero", "Displayed name", "hero_name", "Your full professional name")
add("1.3", "1", "Hero", "Professional positioning statement", "hero_summary", "One concise sentence: expertise + problems solved + impact")
add("1.4", "1", "Hero", "Portrait image", "hero_portrait", "Image filename or final asset URL", "Image")
for item in range(1, 5):
    add(f"1.5.{item}.1", "1", f"Profile credential {item}", "Credential value or headline",
        f"profile_{item:02d}_value", "Example: 4+ years or Supply Chain Strategy")
    add(f"1.5.{item}.2", "1", f"Profile credential {item}", "Credential supporting label",
        f"profile_{item:02d}_label", "Example: Years of experience or Domain expertise")
add("1.6", "1", "Core capabilities", "Section heading", "core_skills_heading", "Example: Core skills & capabilities")
for item in range(1, 7):
    add(f"1.7.{item:02d}.1", "1", f"Core capability {item}", "Capability title",
        f"core_skill_{item:02d}_title", "Short capability name")
    add(f"1.7.{item:02d}.2", "1", f"Core capability {item}", "Capability description",
        f"core_skill_{item:02d}_description", "One sentence explaining the capability")

# Page 2 — Strategy
add("2.1", "2", "Strategy introduction", "Section number and label", "strategy_section_label", "Example: 01 / Strategic traction")
add("2.2", "2", "Strategy introduction", "Main headline", "strategy_heading", "Headline introducing your strategic approach")
add("2.3", "2", "Strategy introduction", "Introductory sentence", "strategy_intro", "One sentence explaining the framework")
for item in range(1, 7):
    add(f"2.4.{item:02d}.1", "2", f"Strategic method {item}", "Method title",
        f"strategy_{item:02d}_title", "Short action-oriented title")
    add(f"2.4.{item:02d}.2", "2", f"Strategic method {item}", "Method description",
        f"strategy_{item:02d}_description", "One sentence describing how this creates value")

# Page 3 — Selected Work
add("3.1", "3", "Work introduction", "Section number and label", "work_section_label", "Example: 02 / Selected work")
add("3.2", "3", "Work introduction", "Main headline", "work_heading", "Headline summarizing the project portfolio")
add("3.3", "3", "Work introduction", "Introductory sentence", "work_intro", "Explain the collection and any confidentiality note")
add("3.4", "3", "Work controls", "Timeline instruction", "work_scroll_label", "Example: Scroll selected work")
for item in range(1, 6):
    number = f"{item:02d}"
    fields = [
        ("1", "Project year or period", "year", "Example: 2025 or 2024–25", "Text"),
        ("2", "Client or industry label", "client", "Masked client or industry name", "Text"),
        ("3", "Project title", "title", "Short case-study title", "Text"),
        ("4", "Project card summary", "summary", "One-sentence mandate and contribution", "Text"),
        ("5", "Primary impact metric", "metric", "Example: ₹X Cr, 20%+, or 50 sites", "Metric"),
        ("6", "Mandate and work detail", "mandate", "Short explanation of context, responsibility and work", "Long text"),
        ("7", "Outcome detail", "outcome", "Short explanation of measurable or strategic result", "Long text"),
    ]
    for suffix, purpose, field, placeholder, value_type in fields:
        add(f"3.5.{number}.{suffix}", "3", f"Selected project {item}", purpose,
            f"work_{number}_{field}", placeholder, value_type)
add("3.6", "3", "Project cards", "Card action label", "work_card_action_label", "Example: View case study")
add("3.7", "3", "Project detail", "Mandate-column heading", "work_detail_mandate_label", "Example: Mandate and my work")
add("3.8", "3", "Project detail", "Outcome-column heading", "work_detail_outcome_label", "Example: What changed")

# Page 4 — About / Resume / Contact
add("4.1", "4", "About introduction", "Section eyebrow", "about_eyebrow", "Example: About me")
add("4.2", "4", "About introduction", "Main headline", "about_heading", "Headline describing what shapes your profile")
tab_names = ["education", "certifications", "skills", "extracurriculars", "interests", "resume", "contact"]
for item, tab in enumerate(tab_names, start=1):
    add(f"4.3.{item}", "4", "About navigation", f"{tab.title()} tab label",
        f"about_tab_{tab}", f"Visible label for the {tab} subsection")
    add(f"4.4.{item}.1", "4", f"About: {tab}", "Subsection heading",
        f"about_{tab}_heading", f"Heading shown when {tab} is selected")
    add(f"4.4.{item}.2", "4", f"About: {tab}", "Subsection introduction or body",
        f"about_{tab}_content", f"Introductory content for {tab}")
for item in range(1, 5):
    add(f"4.5.{item:02d}.1", "4", f"Education entry {item}", "Qualification or degree",
        f"education_{item:02d}_qualification", "Example: MBA")
    add(f"4.5.{item:02d}.2", "4", f"Education entry {item}", "Institution, dates and details",
        f"education_{item:02d}_details", "Institution, period, specialization and score")
add("4.6.1", "4", "Resume", "Resume download action", "resume_download_label", "Example: Download resume")
add("4.7.1", "4", "Contact", "Email address", "contact_email", "Professional email address", "Email")
add("4.7.2", "4", "Contact", "LinkedIn or professional profile", "contact_linkedin", "Full profile URL", "URL")

# Page 5 — Footer
add("5.1", "5", "Footer", "Closing statement", "footer_statement", "Example: Designed with purpose")
add("5.2", "5", "Footer", "Copyright text", "footer_copyright", "Example: © 2026 Your Name")


workbook = Workbook()
sheet = workbook.active
sheet.title = "Content Fields"

headers = list(rows[0].keys())
sheet.append(headers)
for row in rows:
    sheet.append([row[column] for column in headers])

dark = "171512"
accent = "98764D"
warm = "F3EEE7"
paper = "FCFAF6"
line = "D8D0C6"
white = "FFFFFF"
thin = Side(style="thin", color=line)

for cell in sheet[1]:
    cell.fill = PatternFill("solid", fgColor=dark)
    cell.font = Font(color=white, bold=True)
    cell.alignment = Alignment(vertical="center")
sheet.row_dimensions[1].height = 28

page_colors = {
    "0": "EDE7DE",
    "1": "F4EBDD",
    "2": "E8EFEA",
    "3": "E7EDF3",
    "4": "F1E8EF",
    "5": "EEECE8",
}

for row_number in range(2, sheet.max_row + 1):
    page = str(sheet.cell(row_number, 2).value)
    fill = PatternFill("solid", fgColor=page_colors.get(page, paper))
    for column_number in range(1, sheet.max_column + 1):
        cell = sheet.cell(row_number, column_number)
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if column_number in (1, 2, 3):
            cell.fill = fill
    sheet.cell(row_number, 1).font = Font(bold=True, color=accent)
    sheet.cell(row_number, 6).font = Font(name="Menlo", size=10, color=dark)
    sheet.cell(row_number, 9).fill = PatternFill("solid", fgColor="FFFDF9")

status_validation = DataValidation(type="list", formula1='"pending,approved,filled,not applicable"', allow_blank=False)
sheet.add_data_validation(status_validation)
status_validation.add(f"J2:J{sheet.max_row}")

sheet.freeze_panes = "A2"
sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
sheet.sheet_view.showGridLines = False

widths = [14, 8, 24, 26, 31, 32, 48, 14, 42, 17, 32]
for index, width in enumerate(widths, start=1):
    sheet.column_dimensions[get_column_letter(index)].width = width

guide = workbook.create_sheet("Page Guide")
guide.append(["Page", "Page Name", "What belongs here", "ID Pattern"])
for page, (name, description) in PAGES.items():
    guide.append([page, name, description, f"{page}.section.subsection.field"])
for cell in guide[1]:
    cell.fill = PatternFill("solid", fgColor=dark)
    cell.font = Font(color=white, bold=True)
for row in guide.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)
guide.freeze_panes = "A2"
guide.sheet_view.showGridLines = False
for index, width in enumerate([10, 28, 70, 28], start=1):
    guide.column_dimensions[get_column_letter(index)].width = width

instructions = workbook.create_sheet("How to Fill")
instruction_rows = [
    ("How the IDs work", "3.5.02.4 means Page 3 → project collection 5 → project 02 → summary field 4."),
    ("What to edit", "Enter your content only in the 'Your Value' column."),
    ("Status", "Use pending while drafting, approved once wording is agreed, and filled once it is added to the website."),
    ("Technical key", "This stable key connects the spreadsheet row to the website field."),
    ("Do not rename casually", "If a technical key changes, the website mapping and spreadsheet should be updated together."),
  ]
instructions.append(["Topic", "Guidance"])
for topic, guidance in instruction_rows:
    instructions.append([topic, guidance])
for cell in instructions[1]:
    cell.fill = PatternFill("solid", fgColor=dark)
    cell.font = Font(color=white, bold=True)
instructions.column_dimensions["A"].width = 25
instructions.column_dimensions["B"].width = 100
instructions.sheet_view.showGridLines = False
for row in instructions.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
workbook.save(OUTPUT)
print(OUTPUT)
print(f"{len(rows)} fields")
